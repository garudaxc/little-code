import unreal
import csv
import openpyxl
import re
# import functools



def is_valid_gpu_family_name(gpu_family_name):
    if re.match(r'^[-()a-zA-Z0-9.\s]+$', gpu_family_name):
        return True
    
    return False

def ReadGPUFamilyName():
    file_name = r"D:\小工具\device_profile_test\数据表_2024_1118_1050.xlsx"
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Sheet1']
    
    result = []

    for row in sheet.iter_rows(min_row=2, min_col=5, max_col = 5, values_only=True):
        if not is_valid_gpu_family_name(row[0]):
            print(f'invalid gpu family name : {row[0]}' )
            continue
        result.append(row[0])

    result = list(set(result))

    print('number of gpu family names : ', len(result))

    return result


def ReadIosCPUBrand():

    file_name = r"D:\小工具\device_profile_test\apple.xlsx"
    workbook = openpyxl.load_workbook(file_name)
    sheet = workbook['Sheet3']

    ios_cpu_brand = []
    for row in sheet.iter_rows(min_row=4, values_only=True):
        code, grade = row
        code = code.split(':')[0].strip()
        ios_cpu_brand.append(code)

    # ios_cpu_brand = list(set(ios_cpu_brand))
    print('number of ios cpu brand : ', len(ios_cpu_brand))

    return ios_cpu_brand


def Test():
    
    gpu_family_strings = ReadGPUFamilyName()

    profiles = []
    for gpu_family_string in gpu_family_strings:
        result = unreal.K6UtilityFunctionLibrary.test_android_device_profile_match_string(gpu_family_string)
    
        profiles.append({'Name' : gpu_family_string, 'profile' : result[0], 'score' : result[2]})

    profiles.sort(key = lambda x : x['score'], reverse=True)

    print(f'number of device profiles : {len(profiles)}')
    
    for item in profiles:
        print(f'{item["Name"]} : {item["profile"]} : {item["score"]}')



class DeviceMappingItem:
    def __init__(self, mapping_str : str):
        self.match_str, self.device_name = mapping_str.split('=')
        self.re = re.compile(self.match_str + '$')

    def match(self, gpu_family_name):
        return self.re.match(gpu_family_name)


def TestIosDevice():

    cpu_brands = ReadIosCPUBrand()

    IOSDeviceMappings, Profiles = unreal.K6UtilityFunctionLibrary.get_ios_device_profile_names()
    print(f'number of device profiles : {len(Profiles)}, number of device mappings : {len(IOSDeviceMappings)}')

    IOSDeviceMappings = [DeviceMappingItem(item) for item in IOSDeviceMappings]
    # print(IOSDeviceMappings)

    Profiles = list(filter(lambda x : len(x.split(' ')) > 1 and x.split(' ')[1] == 'DeviceProfile', Profiles))
    Profiles = [item.split(' ')[0] for item in Profiles]

    def try_match(cpu_brand):
        matched_profiles = []
        for item in IOSDeviceMappings:
            print('\n')
            if item.match(cpu_brand):
                matched_profiles.append((cpu_brand, item))
                # print(f'{cpu_brand} match {item.match_str}')

        if len(matched_profiles) == 0:
            print(f'{cpu_brand} no match !!')
            return False
        
        if len(matched_profiles) > 1:
            print(f'{cpu_brand} match more than one !!')
            print(matched_profiles)
            return False

        cpu_brand, mapping_item = matched_profiles[0]
        if not (mapping_item.device_name in Profiles):
            print(f'{cpu_brand} match {mapping_item.device_name} but not in profiles')
            return False

        return True

    for cpu in cpu_brands:
        try_match(cpu)



    # device_profiles = unreal.K6UtilityFunctionLibrary.test_ios_device_profile_selection()

    # for item in device_profiles:
    #     print(f'{item[0]} : {item[1]} : {item[2]}')


def TestWindowsDevice():
    samples_file = r'D:\小工具\device_profile_test\pc_device_2025_0310_1142.xlsx'
    workbook = openpyxl.load_workbook(samples_file)
    sheet = workbook['Sheet1']

    gpu_names = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        gpu_family_name = row[2]
        if gpu_family_name is None:
            continue
        if gpu_family_name in gpu_names:
            continue
        gpu_names.append(gpu_family_name)

    print('total gpu names : ', len(gpu_names))

    # for index, gpu_family_name in enumerate(gpu_names):
    #     print(f'{index} : {gpu_family_name}')


    invalid_gpu_names = []
    for gpu_family_name in gpu_names:
        result = unreal.K6UtilityFunctionLibrary.test_windows_devices_profiles(gpu_family_name)
        if result is None:
            invalid_gpu_names.append(gpu_family_name)

    print('invalid gpu names : ', len(invalid_gpu_names))
    for item in invalid_gpu_names:
        print(item)

if __name__ == "__main__":
    # TestIosDevice()
    TestWindowsDevice()
